#!/usr/bin/env python3
"""
Bygger energibesparing_v61.xlsx från daily_summary (API-baserad).
Struktur matchar v53: Sammanfattning + månadsblad med 4 rader/dag.
"""
import sqlite3
import os
from datetime import datetime
from calendar import monthrange
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH = "data/sigen.db"
OUT_PATH = "energibesparing_v61.xlsx"

FONT = "Arial"
NAVY = "FF1F4E79"
BLUE = "FF2E75B6"
SOL_ORANGE = "FFD4960A"
BAT_RED = "FFC00000"
EVDC_PURPLE = "FF7030A0"
GREEN_LIGHT = "FFEBF5EB"
GREEN_FILL = "FFD4EDDA"
BAT_FILL = "FFF8D7DA"
EVDC_FILL = "FFE6D7F0"
BLUE_FILL = "FFD6E4F7"
WHITE = "FFFFFFFF"
GREY = "FF595959"

NUM_KR = '#,##0.00;[Red]\\-#,##0.00;\\-'
NUM_PCT = '0.0%;[Red]\\-0.0%;\\-'

THIN = Side(border_style="thin", color="FF808080")
MEDIUM = Side(border_style="medium", color="FF1F4E79")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_MEDIUM = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)

MONTH_NAMES_SV = {1:"Januari",2:"Februari",3:"Mars",4:"April",5:"Maj",6:"Juni",
                  7:"Juli",8:"Augusti",9:"September",10:"Oktober",11:"November",12:"December"}


def make_font(bold=False, size=10, color="FF000000"):
    return Font(name=FONT, bold=bold, size=size, color=color)


def load_data():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.execute("""
        SELECT date, rows, sol_kwh, last_kwh, import_kwh, export_kwh,
               bat_lad_kwh, bat_url_kwh, evdc_kwh,
               spot_min, spot_max, spot_mean,
               sol_kr, bat_kr, total_kr,
               bat_lad_nat, bat_lad_sol, bat_url_nat, bat_url_last,
               sol_last, sol_nat,
               evdc_url_kwh, evdc_url_last, evdc_url_nat, evdc_kr
        FROM daily_summary
        ORDER BY date
    """)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def format_date_short(date_str):
    _, m, d = date_str.split("-")
    return f"{m}-{d}"


def month_key(date_str):
    y, m, _ = date_str.split("-")
    return (int(y), int(m))


def month_label(year, month, days):
    name = MONTH_NAMES_SV[month]
    _, last = monthrange(year, month)
    ds = sorted(d["date"] for d in days)
    first_d = int(ds[0].split("-")[2])
    last_d = int(ds[-1].split("-")[2])
    partial = (first_d != 1) or (last_d != last)
    hour_res = (year, month) in [(2025, 8), (2025, 9)]
    suffix = ""
    if partial:
        suffix += f" ({first_d}–{last_d} {name.lower()[:3]}) *"
    if hour_res:
        suffix += "†"
    return f"{name} {year}{suffix}"


def build_summary(wb, all_days):
    ws = wb.create_sheet("Sammanfattning", 0)
    
    months = {}
    for d in all_days:
        k = month_key(d["date"])
        months.setdefault(k, []).append(d)
    sorted_months = sorted(months.keys())
    
    first_d = all_days[0]["date"]
    last_d = all_days[-1]["date"]
    fy, fm, _ = first_d.split("-")
    ly, lm, _ = last_d.split("-")
    fmo = MONTH_NAMES_SV[int(fm)].upper()[:3]
    lmo = MONTH_NAMES_SV[int(lm)].upper()[:3]
    
    ws["A1"] = f"ENERGIKOSTNADSBESPARING – SAMMANFATTNING {fmo} {fy}–{lmo} {ly}"
    ws["A1"].font = make_font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", start_color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:P1")
    ws.row_dimensions[1].height = 27.75
    
    ws["A2"] = "Data från Sigen API 5-min (energy_5min_v2)  |  bat_url_evdc @ inkopspris  |  Sol inkl. bat_lad_sol"
    ws["A2"].font = make_font(size=9, color=WHITE)
    ws["A2"].fill = PatternFill("solid", start_color=BLUE)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:P2")
    ws.row_dimensions[2].height = 15.75
    ws.row_dimensions[3].height = 4.5
    
    # Grupprubriker rad 4
    for coord, val, fill, span in [
        ("B4", "☀ SOLCELLER", SOL_ORANGE, "B4:C4"),
        ("D4", "⚡ HEMBATTERI", BAT_RED, "D4:F4"),
        ("G4", "🚗 EVDC", EVDC_PURPLE, "G4:I4"),
    ]:
        ws[coord] = val
        ws[coord].font = make_font(bold=True, size=9, color=WHITE)
        ws[coord].fill = PatternFill("solid", start_color=fill)
        ws[coord].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(span)
    ws.row_dimensions[4].height = 19.5
    
    # Rad 5 - kolumnrubriker
    for coord, val, fill in [
        ("A5", "Månad", NAVY),
        ("B5", "kr", SOL_ORANGE),
        ("C5", "andel", SOL_ORANGE),
        ("D5", "kr\n(netto)", BAT_RED),
        ("E5", "andel", BAT_RED),
        ("F5", "Sol-lad.\navdr.", BAT_RED),
        ("G5", "kr\n(netto)", EVDC_PURPLE),
        ("H5", "andel", EVDC_PURPLE),
        ("I5", "Laddkost.\n(kr)", EVDC_PURPLE),
        ("K5", "Vår\nber.", NAVY),
        ("L5", "Sigenergy", NAVY),
        ("M5", "Avvik.", NAVY),
    ]:
        ws[coord] = val
        ws[coord].font = make_font(bold=True, size=9, color=WHITE)
        ws[coord].fill = PatternFill("solid", start_color=fill)
        ws[coord].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[5].height = 36.0
    
    # Datarader
    row = 6
    for (yr, mo) in sorted_months:
        days = months[(yr, mo)]
        label = month_label(yr, mo, days)
        sol_kr = sum(d["sol_kr"] for d in days)
        bat_kr = sum(d["bat_kr"] for d in days)
        evdc_kr_month = sum(d.get("evdc_kr") or 0 for d in days)
        v2h_kwh_month = sum(d.get("evdc_url_kwh") or 0 for d in days)
        total_kr = sum(d["total_kr"] for d in days)
        f_val = -sum((d.get("bat_lad_sol") or 0) * (d.get("spot_mean") or 0) for d in days)
        
        ws.cell(row=row, column=1, value=label).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", start_color=GREEN_LIGHT)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        ws.cell(row=row, column=2, value=sol_kr).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=2).fill = PatternFill("solid", start_color=GREEN_FILL)
        ws.cell(row=row, column=2).number_format = NUM_KR
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row, column=3, value=(sol_kr/total_kr if total_kr else 0)).font = make_font(size=10)
        ws.cell(row=row, column=3).fill = PatternFill("solid", start_color=GREEN_FILL)
        ws.cell(row=row, column=3).number_format = NUM_PCT
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row, column=4, value=bat_kr).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=4).fill = PatternFill("solid", start_color=BAT_FILL)
        ws.cell(row=row, column=4).number_format = NUM_KR
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row, column=5, value=(bat_kr/total_kr if total_kr else 0)).font = make_font(size=10)
        ws.cell(row=row, column=5).fill = PatternFill("solid", start_color=BAT_FILL)
        ws.cell(row=row, column=5).number_format = NUM_PCT
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row, column=6, value=f_val).font = make_font(size=10, color=GREY)
        ws.cell(row=row, column=6).fill = PatternFill("solid", start_color=BAT_FILL)
        ws.cell(row=row, column=6).number_format = NUM_KR
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="right", vertical="center")
        
        # EVDC - kolumn 7 = kr (V2H netto), 8 = andel, 9 = V2H kWh
        c = ws.cell(row=row, column=7, value=evdc_kr_month)
        c.font = make_font(bold=True, size=10)
        c.fill = PatternFill("solid", start_color=EVDC_FILL)
        c.number_format = NUM_KR
        c.alignment = Alignment(horizontal="right", vertical="center")
        
        c = ws.cell(row=row, column=8, value=(evdc_kr_month/total_kr if total_kr else 0))
        c.font = make_font(size=10)
        c.fill = PatternFill("solid", start_color=EVDC_FILL)
        c.number_format = NUM_PCT
        c.alignment = Alignment(horizontal="right", vertical="center")
        
        c = ws.cell(row=row, column=9, value=v2h_kwh_month)
        c.font = make_font(size=10, color=GREY)
        c.fill = PatternFill("solid", start_color=EVDC_FILL)
        c.number_format = "0.0"
        c.alignment = Alignment(horizontal="right", vertical="center")
        
        ws.cell(row=row, column=11, value=total_kr).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=11).fill = PatternFill("solid", start_color=BLUE_FILL)
        ws.cell(row=row, column=11).number_format = NUM_KR
        ws.cell(row=row, column=11).alignment = Alignment(horizontal="right", vertical="center")
        
        for col in [12, 13]:
            c = ws.cell(row=row, column=col, value="–")
            c.font = make_font(size=10, color=GREY)
            c.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[row].height = 19.5
        row += 1
    
    total_row = row
    ws.cell(row=total_row, column=1, value=f"TOTALT ({fmo.capitalize()} {fy}–{lmo.capitalize()} {ly})")
    ws.cell(row=total_row, column=1).font = make_font(bold=True, size=11, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color=NAVY)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=16)
    ws.row_dimensions[total_row].height = 36.0
    
    info_row = total_row + 1
    ws.cell(row=info_row, column=1, value=(
        "* Partiell månad.  † Timupplöst spot (aug+sep 2025).  "
        "Data från Sigen API 5-min (energy_5min_v2)."
    ))
    ws.cell(row=info_row, column=1).font = make_font(size=9, color=GREY)
    ws.cell(row=info_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=info_row, start_column=1, end_row=info_row, end_column=16)
    ws.row_dimensions[info_row].height = 30
    
    kt_row = info_row + 2
    ws.cell(row=kt_row, column=10, value="Grand total:").font = make_font(bold=True, size=11)
    ws.cell(row=kt_row, column=10).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(row=kt_row, column=11, value=f"=SUM(K6:K{row-1})").font = make_font(bold=True, size=11)
    ws.cell(row=kt_row, column=11).fill = PatternFill("solid", start_color=BLUE_FILL)
    ws.cell(row=kt_row, column=11).number_format = NUM_KR
    ws.cell(row=kt_row, column=11).alignment = Alignment(horizontal="right", vertical="center")
    
    widths = {"A": 28, "B": 12, "C": 8, "D": 11, "E": 8, "F": 10,
              "G": 11, "H": 8, "I": 10, "J": 3, "K": 12, "L": 12,
              "M": 12, "N": 3, "O": 3, "P": 3}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w


def build_month(wb, year, month, days):
    name = MONTH_NAMES_SV[month]
    sheet_name = f"{name[:3]} {year}"
    ws = wb.create_sheet(sheet_name)
    
    ws["A1"] = f"ENERGIKOSTNADSBESPARING – {name.upper()} {year}"
    ws["A1"].font = make_font(bold=True, size=13, color=WHITE)
    ws["A1"].fill = PatternFill("solid", start_color=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A1:S1")
    ws.row_dimensions[1].height = 27.75
    
    if year == 2025:
        subtitle = "Inkop: 1,25×spot+0,9532  |  Forsalj: spot+0,72  |  Data: Sigen API 5-min"
    elif year == 2026 and month < 4:
        subtitle = "Inkop: 1,25×spot+0,935  |  Forsalj: spot+0,104  |  Data: Sigen API 5-min"
    else:
        subtitle = "Inkop: 1,25×(spot+0,604)+0,04  |  Forsalj: spot+0,104  |  Data: Sigen API 5-min"
    
    if (year, month) in [(2025, 8), (2025, 9)]:
        subtitle += "  |  OBS: Timupplöst spot"
    
    ws["A2"] = subtitle
    ws["A2"].font = make_font(size=9, color=WHITE)
    ws["A2"].fill = PatternFill("solid", start_color=BLUE)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("A2:S2")
    ws.row_dimensions[2].height = 15.75
    ws.row_dimensions[3].height = 4.5
    
    headers = [
        ("A", "Dag", NAVY),
        ("B", "Bat lad.\nfrån nät", BAT_RED),
        ("C", "Bat lad.\nfrån sol\n(☀)", BAT_RED),
        ("D", "Bat lad.\nfrån sol\n(⚡avdr)", BAT_RED),
        ("E", "EVDC lad.\nfrån sol\n(☀)", EVDC_PURPLE),
        ("F", "Bat url.\n→ nät", BAT_RED),
        ("G", "Bat url.\n→ last", BAT_RED),
        ("H", "Bat url.\n→ EVDC\n(inkop)", BAT_RED),
        ("I", "EVDC url.\n→ nät", EVDC_PURPLE),
        ("J", "EVDC url.\n→ last", EVDC_PURPLE),
        ("K", "Sol\n→ last", SOL_ORANGE),
        ("L", "Sol\n→ nät", SOL_ORANGE),
        ("M", "EVDC\nladdkost.\n(FIFO)", EVDC_PURPLE),
        ("O", "Total\n(kr)", NAVY),
        ("Q", "☀ Sol\n(kr)", SOL_ORANGE),
        ("R", "⚡ Batteri\n(kr)", BAT_RED),
        ("S", "🚗 EVDC\n(kr)", EVDC_PURPLE),
    ]
    for col, val, fill in headers:
        c = ws[f"{col}4"]
        c.value = val
        c.font = make_font(bold=True, size=9, color=WHITE)
        c.fill = PatternFill("solid", start_color=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_THIN
    ws.row_dimensions[4].height = 42
    
    sorted_days = sorted(days, key=lambda d: d["date"])
    row = 5
    day_rows_for_total = []
    
    for day in sorted_days:
        day_short = format_date_short(day["date"])
        sol_kr = day.get("sol_kr") or 0
        bat_kr = day.get("bat_kr") or 0
        total_kr = day.get("total_kr") or 0
        evdc_kr = day.get("evdc_kr") or 0
        sm = day.get("spot_mean") or 0
        
        b = (day.get("bat_lad_nat") or 0) * (sm + 0.9532)
        c = (day.get("bat_lad_sol") or 0) * (sm + 0.72)
        d_val = -c
        e = 0
        f_val = (day.get("bat_url_nat") or 0) * (sm + 0.72)
        g = (day.get("bat_url_last") or 0) * (sm + 0.9532)
        h = 0
        i_val = 0
        j = 0
        k_val = (day.get("sol_last") or 0) * (sm + 0.9532)
        l_val = (day.get("sol_nat") or 0) * (sm + 0.72)
        m_val = 0
        
        ws.cell(row=row, column=1, value=day_short).font = make_font(bold=True, size=10, color=NAVY)
        ws.cell(row=row, column=1).fill = PatternFill("solid", start_color=WHITE)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = BORDER_THIN
        
        for idx, v in enumerate([b, c, d_val, e, f_val, g, h, i_val, j, k_val, l_val, m_val]):
            cell = ws.cell(row=row, column=idx+2, value=v)
            cell.font = make_font(size=10)
            cell.fill = PatternFill("solid", start_color=WHITE)
            cell.number_format = NUM_KR
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = BORDER_THIN
        
        ws.cell(row=row, column=15, value=total_kr).font = make_font(bold=True, size=10)
        ws.cell(row=row, column=15).fill = PatternFill("solid", start_color=BLUE_FILL)
        ws.cell(row=row, column=15).number_format = NUM_KR
        ws.cell(row=row, column=15).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=15).border = BORDER_THIN
        
        ws.cell(row=row, column=17, value=sol_kr).font = make_font(size=10)
        ws.cell(row=row, column=17).fill = PatternFill("solid", start_color=GREEN_FILL)
        ws.cell(row=row, column=17).number_format = NUM_KR
        ws.cell(row=row, column=17).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=17).border = BORDER_THIN
        
        ws.cell(row=row, column=18, value=bat_kr).font = make_font(size=10)
        ws.cell(row=row, column=18).fill = PatternFill("solid", start_color=BAT_FILL)
        ws.cell(row=row, column=18).number_format = NUM_KR
        ws.cell(row=row, column=18).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=18).border = BORDER_THIN
        
        ws.cell(row=row, column=19, value=evdc_kr).font = make_font(size=10)
        ws.cell(row=row, column=19).fill = PatternFill("solid", start_color=EVDC_FILL)
        ws.cell(row=row, column=19).number_format = NUM_KR
        ws.cell(row=row, column=19).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=19).border = BORDER_THIN
        
        day_rows_for_total.append(row)
        
        for offset, (label, fill) in enumerate([
            ("☀ Solceller", SOL_ORANGE),
            ("⚡ Hembatteri", BAT_RED),
            ("🚗 EVDC", EVDC_PURPLE)
        ]):
            r2 = row + 1 + offset
            ws.cell(row=r2, column=1, value=label).font = make_font(size=8, color=GREY)
            ws.cell(row=r2, column=1).fill = PatternFill("solid", start_color=fill)
            ws.cell(row=r2, column=1).alignment = Alignment(horizontal="left", vertical="center")
            
            val = [sol_kr, bat_kr, evdc_kr][offset]
            fill2 = [GREEN_FILL, BAT_FILL, EVDC_FILL][offset]
            ws.cell(row=r2, column=15, value=val).font = make_font(size=9)
            ws.cell(row=r2, column=15).fill = PatternFill("solid", start_color=fill2)
            ws.cell(row=r2, column=15).number_format = NUM_KR
            ws.cell(row=r2, column=15).alignment = Alignment(horizontal="right", vertical="center")
        
        row += 4
    
    total_row = row
    ws.cell(row=total_row, column=1, value="TOTALT").font = make_font(bold=True, size=10, color=WHITE)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", start_color=NAVY)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=total_row, column=1).border = BORDER_MEDIUM
    
    for col_letter in ["B","C","D","E","F","G","H","I","J","K","L","M","O","Q","R","S"]:
        cells = "+".join(f"{col_letter}{r}" for r in day_rows_for_total)
        cell = ws.cell(row=total_row, column=ord(col_letter)-ord('A')+1, value=f"={cells}")
        cell.font = make_font(bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", start_color=NAVY)
        cell.number_format = NUM_KR
        cell.alignment = Alignment(horizontal="right", vertical="center")
        cell.border = BORDER_MEDIUM
    ws.row_dimensions[total_row].height = 22
    
    widths = {"A": 12, "B": 12, "C": 12, "D": 12, "E": 12, "F": 11,
              "G": 11, "H": 12, "I": 11, "J": 11, "K": 10, "L": 10,
              "M": 12, "N": 1, "O": 13, "P": 1, "Q": 12, "R": 12, "S": 12}
    for c, w in widths.items():
        ws.column_dimensions[c].width = w


def main():
    days = load_data()
    print(f"Loaded {len(days)} days from DB")
    print(f"Range: {days[0]['date']} → {days[-1]['date']}")
    print(f"Grand total: {sum(d['total_kr'] for d in days):.2f} kr")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    build_summary(wb, days)
    
    months = {}
    for d in days:
        k = month_key(d["date"])
        months.setdefault(k, []).append(d)
    
    for (yr, mo) in sorted(months.keys()):
        build_month(wb, yr, mo, months[(yr, mo)])
    
    wb.save(OUT_PATH)
    print(f"\n✓ Saved: {OUT_PATH}")
    print(f"  Sheets: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")


if __name__ == "__main__":
    main()
