"""Processes downloaded Sigen ZIP files: extract XLSX, parse, save to SQLite."""

import os
import sys
import zipfile
import logging
from pathlib import Path
from datetime import datetime

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))
from db import connect

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("process")

DOWNLOAD_DIR = Path(__file__).parent.parent / "data" / "sigen_downloads"
PROCESSED_DIR = DOWNLOAD_DIR / "processed"
PROCESSED_DIR.mkdir(parents=True, exist_ok=True)


def process_zip(zip_path: Path):
    """Extract XLSX from ZIP and insert rows into energy_5min."""
    log.info("Processing %s", zip_path.name)

    rows_inserted = 0
    rows_updated = 0

    with zipfile.ZipFile(zip_path, "r") as zf:
        xlsx_names = [n for n in zf.namelist() if n.lower().endswith(".xlsx")]
        if not xlsx_names:
            log.warning("No XLSX inside ZIP, skipping")
            return 0

        for xlsx_name in xlsx_names:
            log.info("  Reading %s", xlsx_name)
            with zf.open(xlsx_name) as xf:
                wb = openpyxl.load_workbook(xf, data_only=True)
                ws = wb[wb.sheetnames[0]]

                # Check header to identify format
                header = [c.value for c in ws[1]]
                log.info("  Header: %s", header)

                # Expected: Datum | Last (kW) | Bat (kW) | EVDC (kW) | Grid (kW) | PV (kW)
                if len(header) < 6:
                    log.warning("  Unexpected column count: %d", len(header))
                    continue

                with connect() as conn:
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or row[0] is None:
                            continue
                        try:
                            ts_str = row[0]
                            if isinstance(ts_str, datetime):
                                ts = ts_str
                            else:
                                ts = datetime.strptime(str(ts_str), "%Y-%m-%d %H:%M:%S")
                        except (ValueError, TypeError):
                            continue

                        ts_local = ts.isoformat(timespec="seconds")
                        date_local = ts.date().isoformat()

                        def f(v):
                            try:
                                return float(v) if v is not None else 0.0
                            except (ValueError, TypeError):
                                return 0.0

                        last_kw = f(row[1])
                        bat_kw = f(row[2])
                        evdc_kw = f(row[3])
                        grid_kw = f(row[4])
                        pv3_kw = f(row[5])

                        cur = conn.execute(
                            "INSERT OR REPLACE INTO energy_5min "
                            "(ts_local, date_local, last_kw, bat_kw, evdc_kw, grid_kw, pv3_kw) "
                            "VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (ts_local, date_local, last_kw, bat_kw, evdc_kw, grid_kw, pv3_kw)
                        )
                        if cur.rowcount > 0:
                            rows_inserted += 1

    log.info("  Inserted/updated %d rows", rows_inserted)

    # Move to processed
    target = PROCESSED_DIR / zip_path.name
    zip_path.rename(target)
    log.info("  Moved to %s", target.name)

    return rows_inserted


def main():
    zips = sorted(DOWNLOAD_DIR.glob("*.zip"))
    log.info("Found %d ZIP files to process", len(zips))

    total = 0
    for zip_path in zips:
        try:
            total += process_zip(zip_path)
        except Exception as e:
            log.error("Failed to process %s: %s", zip_path.name, e)

    log.info("Total rows processed: %d", total)
    return 0


if __name__ == "__main__":
    sys.exit(main())
